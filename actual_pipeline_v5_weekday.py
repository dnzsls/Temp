# =============================================================================
# AYLIK ÇALIŞTIRMA V9 + UYUM KONTROLÜ
# =============================================================================
# Bir ayın TÜM günlerini koşar:
#   - Haftaiçi → actual_pipeline_v9_weekly.run_week_all_queues
#     (Pzt-Cum bloklarına ayırıp HER HAFTAYI tek MIP'te çözer)
#     V9 = V8 fallback ladder + Stage 4 (coverage shortfall)
#   - Haftasonu → weekend_model_gelistirim_final.run_all_queues (günlük MIP)
#
# V9 farkları:
#   - Stage 4 (coverage shortfall) sayesinde kadro yetmediğinde de çözüm üretir
#   - Her gün/queue için solution_stage notu çıkarılır (orijinal / shrinkage -X /
#     shortfall N kişi-slot)
#   - HÜCRE 7 sade çıktı: gün başına kuyruk bazlı OK [not] / INFEASIBLE
#   - HÜCRE 9/10/11 — KİTLE / GOLD / KURUMSAL plan özetleri ayrı hücrelerde
#
# Sonuç: results = {date_str: {queue: pipeline_output} | None}
# Cache: monthly_v9_YYYY_MM.pkl
#
# DOSYALAR:
#   - actual_pipeline_v9_weekly.py     → haftaiçi MIP motoru (V9, Stage 4 dahil)
#   - weekend_model_gelistirim_final.py → haftasonu MIP motoru (günlük)
#
# Hücre hücre çalıştır (# %% [HÜCRE N] ile işaretli).
# =============================================================================


# %% [HÜCRE 1] — Importlar
import io
import os
import pickle
import calendar
import contextlib
import pandas as pd

# V9 weekly — haftaiçi 5 günlük blok için (Stage 4 coverage shortfall ile)
from actual_pipeline_v9_weekly import (
    load_aht_from_df,
    run_week_all_queues,
    prepare_calls_30,
)
# weekend pipeline — haftasonu günlük
from weekend_model_gelistirim_final import (
    run_all_queues as run_weekend_all,
)


# %% [HÜCRE 2] — Veri çekme
# Mevcut notebook'taki veri çekme kodunu BURAYA YAPIŞTIR.
# Gereken:
#   - df_calls (en az tüm ayı içermeli)
#   - df_aht
#   - df_actual (gerçek vardiya verisi)
#   - df_shifts_dict = {'kitle': df, 'kurumsal': df, 'gold': df}


# %% [HÜCRE 3] — CONFIG_WEEKDAY
# run_weekly_v9.py'deki CONFIG_WEEKDAY'i buraya kopyala.
# V9 config eklemeleri (mip altında her queue için):
#   'coverage_shortfall': {'enabled': True, 'penalty': 1000.0}
#   'weekly_min_per_shift_fallback': {'enabled': False, ...}  (min sert kısıt)


# %% [HÜCRE 4] — CONFIG_WEEKEND


# %% [HÜCRE 5] — AHT yükle (her iki config için)
CONFIG_WEEKDAY['sub_queues'] = load_aht_from_df(df_aht, config=CONFIG_WEEKDAY)
CONFIG_WEEKEND['sub_queues'] = load_aht_from_df(df_aht, config=CONFIG_WEEKEND)


# %% [HÜCRE 6] — Aylık çalıştırma fonksiyonu
# Self-contained: notebook'ta HÜCRE 1 tekrar koşturulmasa bile çalışsın diye
# kritik import'ları burada da güvene alıyoruz.
import io
import contextlib

DAY_LABELS_EN = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
DAY_LABELS_TR = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Pzr']


def _wrap_weekly_info_as_daily(info, df_calls_30, queue, date_str):
    """v9 weekly info_per_day[d] çıktısını v6 daily benzeri dict'e çevirir."""
    d = pd.to_datetime(date_str)
    calls_col = f"{queue}_total"
    df_day = df_calls_30[df_calls_30['data_date'] == d]
    calls_by_slot = (dict(zip(df_day['slot_30'], df_day[calls_col]))
                     if calls_col in df_day.columns else {})
    return {
        'mip_info': info,
        'erlang_by_slot': info.get('erlang_by_slot', {}),
        'calls_by_slot': calls_by_slot,
        'actual': None,
        'date': d,
        'queue': queue,
    }


def _short_stage_note(solution_stage, total_shortfall_week=0):
    """V9 — solution_stage string'ini kısa, okunaklı bir nota çevirir.

    Örnekler:
      'STAGE 1: original'                       → 'orijinal'
      'STAGE 2 (per-day): Mon=-0.10'            → 'shrinkage Mon=-0.10'
      'STAGE 2 (uniform): all -0.20'            → 'shrinkage all -0.20'
      'STAGE 3 (per-day): Mon=12'               → 'min_per_shift Mon=12'
      'STAGE 4: coverage_shortfall'             → 'shortfall N kişi-slot'
    """
    if not solution_stage:
        return '?'
    s = solution_stage
    if s.startswith('STAGE 1'):
        return 'orijinal'
    if s.startswith('STAGE 2'):
        # 'STAGE 2 (per-day): Mon=-0.10, Tue=-0.05' veya 'STAGE 2 (uniform): all -0.20'
        try:
            payload = s.split(': ', 1)[1]
        except IndexError:
            payload = s
        return f"shrinkage {payload}"
    if s.startswith('STAGE 3'):
        try:
            payload = s.split(': ', 1)[1]
        except IndexError:
            payload = s
        return f"min_per_shift {payload}"
    if s.startswith('STAGE 4'):
        return f"shortfall {total_shortfall_week} kişi-slot"
    return s


def _run_weekend_silent(df_calls, df_actual, df_shifts_dict, date_str, config):
    """Haftasonu pipeline'ı stdout'u yutarak çağır — sadece sonuç dict'i alıyoruz."""
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        r = run_weekend_all(df_calls, df_actual, df_shifts_dict,
                            date_str, config=config)
    return r


def run_month(year, month, df_calls, df_actual, df_shifts_dict,
              cfg_weekday, cfg_weekend, queues=('kitle', 'kurumsal', 'gold'),
              verbose=False):
    """Bir ayın tüm günlerini koşar, results dict döner.

    verbose=False (default): per-day kuyruk bazlı tek satır OK[not] / INFEASIBLE
    verbose=True: V9 pipeline'ın tüm verbose çıktısı (debug için)

    Returns: {date_str: {queue: {mip_info, ...}} | None}
    """
    results = {}
    cal = calendar.Calendar()

    all_days = [d for d in cal.itermonthdates(year, month) if d.month == month]

    # Haftaiçi blokları (ardışık Pzt-Cum)
    weekday_blocks = []
    current = []
    for d in all_days:
        if d.weekday() < 5:
            current.append(d)
        else:
            if current:
                weekday_blocks.append(current)
                current = []
    if current:
        weekday_blocks.append(current)

    df_calls_30 = prepare_calls_30(df_calls, config=cfg_weekday)

    # ---- HAFTAİÇİ BLOKLARI ----
    for block in weekday_blocks:
        block_dates = [d.strftime('%Y-%m-%d') for d in block]
        block_label = f"{block_dates[0]} → {block_dates[-1]} ({len(block)} gün)"
        print(f"\n▶ Haftaiçi blok: {block_label}")

        try:
            if verbose:
                week_results = run_week_all_queues(
                    df_calls=df_calls,
                    df_shifts_by_queue=df_shifts_dict,
                    target_dates=block_dates,
                    config=cfg_weekday,
                    queues=queues,
                    verbose=True,
                )
            else:
                # Pipeline detaylarını yutarak çağır — yalnız result lazım
                buf = io.StringIO()
                with contextlib.redirect_stdout(buf):
                    week_results = run_week_all_queues(
                        df_calls=df_calls,
                        df_shifts_by_queue=df_shifts_dict,
                        target_dates=block_dates,
                        config=cfg_weekday,
                        queues=queues,
                        verbose=False,
                    )
        except Exception as e:
            print(f"  ✗ blok hatası: {e}")
            for d in block:
                results[d.strftime('%Y-%m-%d')] = None
            continue

        # Her gün için: kuyruk bazlı OK [not] / INFEASIBLE
        for d in block:
            date_str = d.strftime('%Y-%m-%d')
            day_label = DAY_LABELS_EN[d.weekday()]
            day_name = DAY_LABELS_TR[d.weekday()]

            day_result = {}
            parts = []
            for queue in queues:
                q_res = week_results.get(queue)
                if q_res is None:
                    parts.append(f"{queue}: INFEASIBLE")
                    continue
                info = q_res.get('info_per_day', {}).get(day_label)
                if info is None:
                    parts.append(f"{queue}: INFEASIBLE")
                    continue
                day_result[queue] = _wrap_weekly_info_as_daily(
                    info, df_calls_30, queue, date_str)
                note = _short_stage_note(
                    q_res.get('solution_stage'),
                    q_res.get('total_shortfall_week', 0),
                )
                parts.append(f"{queue}: OK [{note}]")

            if not day_result:
                results[date_str] = None
                print(f"  ✗ {date_str} ({day_name}) — tümü INFEASIBLE")
            else:
                results[date_str] = day_result
                # Shortfall varsa uyarı sembolü
                has_sf = any(
                    week_results.get(q, {}).get('total_shortfall_week', 0) > 0
                    for q in day_result
                )
                mark = "⚠" if has_sf else "✓"
                print(f"  {mark} {date_str} ({day_name}) — " + "  |  ".join(parts))

    # ---- HAFTASONU GÜNLERİ ----
    weekend_days = [d for d in all_days if d.weekday() >= 5]
    if weekend_days:
        print(f"\n▶ Haftasonu günleri: {len(weekend_days)} gün")
    for d in weekend_days:
        date_str = d.strftime('%Y-%m-%d')
        day_name = DAY_LABELS_TR[d.weekday()]
        try:
            if verbose:
                r = run_weekend_all(df_calls, df_actual, df_shifts_dict,
                                    date_str, config=cfg_weekend)
            else:
                r = _run_weekend_silent(df_calls, df_actual, df_shifts_dict,
                                        date_str, cfg_weekend)
            results[date_str] = r
            if r:
                parts = []
                for queue in queues:
                    if queue in r and r[queue] is not None:
                        parts.append(f"{queue}: OK")
                    else:
                        parts.append(f"{queue}: INFEASIBLE")
                print(f"  ✓ {date_str} ({day_name}) — " + "  |  ".join(parts))
            else:
                print(f"  ✗ {date_str} ({day_name}) — boş sonuç")
        except Exception as e:
            print(f"  ✗ {date_str} ({day_name}) — weekend hatası: {e}")
            results[date_str] = None

    return results


# %% [HÜCRE 7] — Aylık çalıştır + cache  (V9 — sade çıktı)
YEAR = 2026
MONTH = 2
CACHE_FILE = f"monthly_v9_{YEAR}_{MONTH:02d}.pkl"

if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, 'rb') as f:
        results = pickle.load(f)
    print(f"Cache'ten yüklendi: {len(results)} gün ({CACHE_FILE})")
else:
    # verbose=False → pipeline detayları yutulur, sadece gün/kuyruk özetleri çıkar
    results = run_month(YEAR, MONTH, df_calls, df_actual, df_shifts_dict,
                        CONFIG_WEEKDAY, CONFIG_WEEKEND, verbose=False)
    with open(CACHE_FILE, 'wb') as f:
        pickle.dump(results, f)
    print(f"\nKaydedildi: {CACHE_FILE}")


# %% [HÜCRE 8] — Haftasonu İnhouse Bütçe Kontrolü (part-time HARİÇ)
WEEKEND_BUDGET = {
    'kitle':    1200,
    'gold':     300,
    'kurumsal': 120,
}

weekend_days = []
for date_str in sorted(results.keys()):
    d = pd.to_datetime(date_str)
    if d.weekday() >= 5:
        weekend_days.append((date_str, d.weekday()))

print(f"\n{'='*100}")
print(f"HAFTA SONU İNHOUSE BÜTÇE KONTROLÜ — {YEAR}/{MONTH:02d}  (part-time HARİÇ)")
print(f"{'='*100}")

print(f"{'Kuyruk':<10} ", end='')
for date_str, wd in weekend_days:
    label = ('Cmt' if wd == 5 else 'Pzr') + date_str[-2:]
    print(f"{label:>7}", end=' ')
print(f"{'Toplam':>8} {'Bütçe':>7} {'Durum':>12}")
print('-' * 100)

for queue, budget in WEEKEND_BUDGET.items():
    inhouse_per_day = []
    for date_str, _ in weekend_days:
        r = results.get(date_str)
        if r and queue in r and r[queue] is not None:
            mi = r[queue]['mip_info']
            inhouse_ft = mi.get('total_inhouse_kisi', 0) - mi.get('total_part_time_kisi', 0)
            inhouse_ft = max(0, inhouse_ft)
        else:
            inhouse_ft = 0
        inhouse_per_day.append(inhouse_ft)

    total = sum(inhouse_per_day)
    diff = total - budget
    if diff > 0:
        status = f"+{diff} AŞIM ✗"
    else:
        status = f"{diff:+d} OK ✓"

    print(f"{queue:<10} ", end='')
    for v in inhouse_per_day:
        print(f"{v:>7}", end=' ')
    print(f"{total:>8} {budget:>7} {status:>12}")

print('-' * 100)
print("Not: total_inhouse_kisi'den total_part_time_kisi çıkarılarak hesaplandı.")


# =============================================================================
# Plan özet + vardiya planı — ortak fonksiyon
# (HÜCRE 9/10/11 bunu kuyruk bazında çağırır)
# =============================================================================

DAY_NAMES = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Pzr']


def print_queue_monthly_plan(plan_queue, results, year, month):
    """Tek kuyruk için aylık plan — gün özet metrikler + vardiya roster.

    V9 farkı: solution_stage ve shortfall bilgisini de gösterir.
    """
    print(f"\n{'='*120}")
    print(f"GÜN GÜN PLAN ÖZETİ — {plan_queue.upper()} — {year}/{month:02d}")
    print(f"{'='*120}")
    print(f"  {'Tarih':<11} {'Gün':<4} {'Çağrı':>7} {'Erl_Pk':>7} {'Erl_T':>7} "
          f"{'MIP_Pk':>7} {'In':>5} {'Out':>5} {'PT':>4} {'Toplam':>7} {'Out%':>5} "
          f"{'Eksik':>6}  {'Çözüm':<28}")
    print('  ' + '-' * 118)

    t_cagri = t_erl_t = 0
    t_in = t_out = t_pt = 0
    t_sf = 0
    days_ok = 0

    for date_str in sorted(results.keys()):
        r = results.get(date_str)
        d = pd.to_datetime(date_str)
        gun = DAY_NAMES[d.weekday()]

        if r is None or plan_queue not in r or r[plan_queue] is None:
            print(f"  {date_str:<11} {gun:<4} {'-':>7} {'-':>7} {'-':>7} "
                  f"{'-':>7} {'-':>5} {'-':>5} {'-':>4} {'-':>7} {'-':>5} "
                  f"{'-':>6}  INFEASIBLE")
            continue

        qr = r[plan_queue]
        mi = qr['mip_info']

        calls_by_slot = qr.get('calls_by_slot', {}) or {}
        cagri = int(sum(calls_by_slot.values()))

        erlang_by_slot = qr.get('erlang_by_slot', {}) or {}
        erl_pk = max(erlang_by_slot.values()) if erlang_by_slot else 0
        erl_t = sum(erlang_by_slot.values())

        mip_by_slot = mi.get('mip_by_slot', {}) or {}
        mip_pk = max(mip_by_slot.values()) if mip_by_slot else 0

        in_ft = max(0, mi.get('total_inhouse_kisi', 0) - mi.get('total_part_time_kisi', 0))
        out = mi.get('total_outsource_kisi', 0)
        pt = mi.get('total_part_time_kisi', 0)
        toplam = in_ft + out + pt
        out_pct = out / toplam if toplam > 0 else 0

        # V9: solution_stage notu + shortfall
        sol_stage = mi.get('solution_stage')
        sf_total = mi.get('total_shortfall', 0)
        note = _short_stage_note(sol_stage, sf_total) if sol_stage else 'weekend'
        sf_str = str(sf_total) if sf_total > 0 else '-'

        print(f"  {date_str:<11} {gun:<4} {cagri:>7,} {erl_pk:>7} {erl_t:>7} "
              f"{mip_pk:>7} {in_ft:>5} {out:>5} {pt:>4} {toplam:>7} {out_pct:>4.0%} "
              f"{sf_str:>6}  {note[:28]:<28}")

        t_cagri += cagri
        t_erl_t += erl_t
        t_in += in_ft
        t_out += out
        t_pt += pt
        t_sf += sf_total
        days_ok += 1

    t_toplam = t_in + t_out + t_pt
    t_out_pct = t_out / t_toplam if t_toplam > 0 else 0
    print('  ' + '-' * 118)
    print(f"  {'AYLIK':<11} {'-':<4} {t_cagri:>7,} {'-':>7} {t_erl_t:>7} "
          f"{'-':>7} {t_in:>5} {t_out:>5} {t_pt:>4} {t_toplam:>7} {t_out_pct:>4.0%} "
          f"{t_sf:>6}")
    print(f"\n  Erl_Pk=eşzamanlı Erlang peak  |  Erl_T=günlük Erlang toplamı")
    print(f"  Eksik=Stage 4 coverage shortfall (kişi-slot)  |  Çözüm=fallback aşaması")
    print(f"  Başarılı: {days_ok}/{len(results)} gün")

    # ---- Vardiya planı (roster) ----
    print(f"\n{'='*100}")
    print(f"GÜN GÜN VARDIYA PLANI — {plan_queue.upper()} — {year}/{month:02d}")
    print(f"{'='*100}")

    for date_str in sorted(results.keys()):
        r = results.get(date_str)
        d = pd.to_datetime(date_str)
        gun = DAY_NAMES[d.weekday()]

        if r is None or plan_queue not in r or r[plan_queue] is None:
            print(f"\n--- {date_str} ({gun}) ---  INFEASIBLE")
            continue

        qr = r[plan_queue]
        mi = qr['mip_info']
        assignments = mi.get('assignments', {})
        sc = mi.get('shift_coverage', {})

        in_ft = max(0, mi.get('total_inhouse_kisi', 0) - mi.get('total_part_time_kisi', 0))
        out = mi.get('total_outsource_kisi', 0)
        pt = mi.get('total_part_time_kisi', 0)
        toplam = in_ft + out + pt
        sf_total = mi.get('total_shortfall', 0)

        header = f"--- {date_str} ({gun}) ---  In={in_ft}  Out={out}  PT={pt}  Toplam={toplam}"
        if sf_total > 0:
            header += f"  ⚠ Eksik={sf_total} kişi-slot"
        print(f"\n{header}")

        active = [(s, c) for s, c in assignments.items() if c > 0]
        if not active:
            print(f"    (vardiya yok)")
            continue

        active.sort(key=lambda x: (sc.get(x[0], {}).get('start', '99:99'),
                                    sc.get(x[0], {}).get('company', 'zzz')))

        # V9: shortfall önerileri varsa kolon ekle
        sf_recs = mi.get('shortfall_recommendations', {}) or {}
        has_rec = bool(sf_recs)

        if has_rec:
            print(f"    {'Vardiya':<32} {'Saat':<14} {'Şirket':<10} "
                  f"{'Kişi':>5} {'+Öneri':>7} {'Önerilen':>9}")
            print(f"    {'-'*84}")
        else:
            print(f"    {'Vardiya':<32} {'Saat':<14} {'Şirket':<10} {'Kişi':>5}")
            print(f"    {'-'*68}")

        for s, cnt in active:
            info = sc.get(s, {})
            saat = f"{info.get('start','--:--')}-{info.get('end','--:--')}"
            comp = info.get('company', '-')
            if has_rec:
                rec = sf_recs.get(s, 0)
                rec_str = f"+{rec}" if rec > 0 else "-"
                new_total = cnt + rec
                mark = " ←" if rec > 0 else "  "
                print(f"    {s:<32} {saat:<14} {comp:<10} "
                      f"{cnt:>5} {rec_str:>7} {new_total:>8}{mark}")
            else:
                print(f"    {s:<32} {saat:<14} {comp:<10} {cnt:>5}")

        if has_rec:
            rec_total = sum(sf_recs.values())
            print(f"    → +{rec_total} kişi eklenirse {sf_total} kişi-slot eksik kapanır")


# %% [HÜCRE 9] — KİTLE — gün gün plan özeti + vardiya planı
print_queue_monthly_plan('kitle', results, YEAR, MONTH)


# %% [HÜCRE 10] — GOLD — gün gün plan özeti + vardiya planı
print_queue_monthly_plan('gold', results, YEAR, MONTH)


# %% [HÜCRE 11] — KURUMSAL — gün gün plan özeti + vardiya planı
print_queue_monthly_plan('kurumsal', results, YEAR, MONTH)


# %% [HÜCRE 12] — Excel'e aktar: vardiya planı (timetable) + haftasonu bütçe
# Layout (referans Excel'e göre):
#   Sheet 1 "Vardiya Planı":
#     - Her hafta için 7 gün YAN YANA (Pzt..Pzr), her gün 8 kolonluk blok:
#         SHIFT | kitle_inh | kitle_out | gold | kurumsal | INH Total | OS Total | Total
#     - Üst satır: gün adı + tarih (her bloğun üstünde)
#     - Satırlar: o haftada aktif tüm vardiya saat aralıkları (start-end)
#     - Bottom: Toplam satırı (her gün için kolon toplamı)
#     - Haftalar üst üste (1. hafta, sonra 2. hafta, vs.)
#   Sheet 2 "Bütçe": haftasonu inhouse bütçe kontrolü (tek tablo)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def export_monthly_plan_to_excel(results, year, month, weekend_budget,
                                  queues=('kitle', 'gold', 'kurumsal'),
                                  output_path=None):
    """Aylık vardiya planını + haftasonu bütçesini Excel'e yazar.

    Sheet 1: 7 gün yan yana, her gün 8 kolon (SHIFT, kitle_inh, kitle_out,
    gold, kurumsal, INH Total, OS Total, Total). Haftalar üst üste.
    """
    if output_path is None:
        ay_isimleri = ['Ocak', 'Subat', 'Mart', 'Nisan', 'Mayis', 'Haziran',
                       'Temmuz', 'Agustos', 'Eylul', 'Ekim', 'Kasim', 'Aralik']
        ay_adi = ay_isimleri[month - 1]
        output_path = f"vardiya_plani_{year}_{month:02d}_{ay_adi}.xlsx"

    all_dates = sorted(results.keys())
    if not all_dates:
        print("⚠ results boş, export atlandı")
        return None

    # Ayı kapsayan Pzt-Pzr haftaları
    first = pd.to_datetime(all_dates[0])
    last = pd.to_datetime(all_dates[-1])
    start_mon = first - pd.Timedelta(days=first.weekday())
    end_sun = last + pd.Timedelta(days=(6 - last.weekday()))

    weeks = []
    cur = start_mon
    while cur <= end_sun:
        week = [(cur + pd.Timedelta(days=i)).strftime('%Y-%m-%d')
                for i in range(7)]
        weeks.append(week)
        cur += pd.Timedelta(days=7)

    DAY_TR = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Pzr']
    COLS_PER_DAY = 8
    SUB_HEADERS = ['SHIFT', 'kitle inhouse', 'kitle outsource', 'gold',
                   'kurumsal', 'INH Total', 'OS Total', 'Total']

    # ---- Workbook + stiller ----
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vardiya Planı'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78',
                              fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    date_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2',
                            fill_type='solid')
    date_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6',
                             fill_type='solid')
    total_font = Font(bold=True)
    thin = Side(border_style='thin', color='888888')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # PT section yerleşimi (Pzr bloğunun sağında, 1 col boşluk sonra)
    PT_COL_GAP = 1
    PT_SECTION_COLS = 3   # SHIFT | Cmt | Pzr
    PT_SECTION_START = 7 * COLS_PER_DAY + PT_COL_GAP + 1   # 1-indexed
    PT_HEADERS = ['SHIFT', 'Cmt', 'Pzr']

    def _collect_week_data(week):
        """Bu hafta için: shift'leri BAŞLANGIÇ saatine göre grupla (PT hariç).

        Outsource'da aynı start farklı end varsa tek satırda toplanır.
        Label = "start-end" (end: önce inhouse'un end'i, yoksa outsource'un).
        Returns:
          - sorted_starts: list of start times
          - labels: {start: "start-end"}
          - counts: counts[start][day_idx][queue][company] = sum
        """
        all_starts = set()
        counts = {}
        ends_per_start_company = {}   # start → company → set of ends

        for day_idx, ds in enumerate(week):
            r = results.get(ds)
            if not r:
                continue
            for queue in queues:
                qr = r.get(queue)
                if not qr:
                    continue
                mi = qr.get('mip_info', {})
                sc = mi.get('shift_coverage', {})
                for s, c in mi.get('assignments', {}).items():
                    if c <= 0:
                        continue
                    sci = sc.get(s, {})
                    company = sci.get('company', 'inhouse')
                    if company == 'part_time':
                        continue   # PT ayrı bloka gider
                    start = sci.get('start', '?')
                    end = sci.get('end', '?')
                    all_starts.add(start)
                    ends_per_start_company.setdefault(start, {})\
                        .setdefault(company, set()).add(end)
                    counts.setdefault(start, {}).setdefault(day_idx, {})\
                        .setdefault(queue, {})
                    counts[start][day_idx][queue][company] = \
                        counts[start][day_idx][queue].get(company, 0) + c

        # Label: inhouse end öncelikli (daha tutarlı), yoksa outsource shortest
        labels = {}
        for start in all_starts:
            ebc = ends_per_start_company.get(start, {})
            if ebc.get('inhouse'):
                end_repr = sorted(ebc['inhouse'])[0]
            elif ebc.get('outsource'):
                end_repr = sorted(ebc['outsource'])[0]
            else:
                end_repr = '?'
            labels[start] = f"{start}-{end_repr}"

        return sorted(all_starts), labels, counts

    def _collect_pt_weekend(week):
        """Sadece Cmt (index 5) ve Pzr (index 6) için part-time atamaları topla.
        Returns: (sorted_pt_shifts [(start, end), ...], pt_counts[(start,end)][day_idx])
        """
        pt_shifts_set = set()
        pt_counts = {}
        for day_idx in (5, 6):
            ds = week[day_idx]
            r = results.get(ds)
            if not r:
                continue
            for queue in queues:
                qr = r.get(queue)
                if not qr:
                    continue
                mi = qr.get('mip_info', {})
                sc = mi.get('shift_coverage', {})
                for s, c in mi.get('assignments', {}).items():
                    if c <= 0:
                        continue
                    sci = sc.get(s, {})
                    if sci.get('company') != 'part_time':
                        continue
                    start = sci.get('start', '?')
                    end = sci.get('end', '?')
                    key = (start, end)
                    pt_shifts_set.add(key)
                    pt_counts.setdefault(key, {})
                    pt_counts[key][day_idx] = pt_counts[key].get(day_idx, 0) + c
        return sorted(pt_shifts_set), pt_counts

    def _row_values(start_counts):
        """Bir start için 7 değer: kitle_inh, kitle_out, gold, kurumsal, INH, OS, Total."""
        kitle_inh = start_counts.get('kitle', {}).get('inhouse', 0)
        kitle_out = start_counts.get('kitle', {}).get('outsource', 0)
        gold_v = sum(start_counts.get('gold', {}).values())
        kurumsal_v = sum(start_counts.get('kurumsal', {}).values())
        inh_total = kitle_inh + gold_v + kurumsal_v
        os_total = kitle_out
        total = inh_total + os_total
        return [kitle_inh, kitle_out, gold_v, kurumsal_v,
                inh_total, os_total, total]

    current_row = 1
    for week in weeks:
        sorted_starts, labels, counts = _collect_week_data(week)
        pt_sorted, pt_counts = _collect_pt_weekend(week)

        if not sorted_starts and not pt_sorted:
            continue

        # 1) Tarih başlık satırı
        for day_idx, ds in enumerate(week):
            col_offset = day_idx * COLS_PER_DAY
            d = pd.to_datetime(ds)
            date_label = f"{DAY_TR[day_idx]} {d.strftime('%d.%m.%Y')}"
            cell = ws.cell(row=current_row, column=col_offset + 1, value=date_label)
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = center
            ws.merge_cells(start_row=current_row, start_column=col_offset + 1,
                           end_row=current_row, end_column=col_offset + COLS_PER_DAY)
        if pt_sorted:
            cell = ws.cell(row=current_row, column=PT_SECTION_START,
                           value="KISMİ ÇALIŞANLAR")
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = center
            ws.merge_cells(
                start_row=current_row, start_column=PT_SECTION_START,
                end_row=current_row,
                end_column=PT_SECTION_START + PT_SECTION_COLS - 1)
        current_row += 1

        # 2) Sub-header
        for day_idx in range(7):
            col_offset = day_idx * COLS_PER_DAY
            for j, h in enumerate(SUB_HEADERS):
                cell = ws.cell(row=current_row, column=col_offset + j + 1, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
        if pt_sorted:
            for j, h in enumerate(PT_HEADERS):
                cell = ws.cell(row=current_row, column=PT_SECTION_START + j, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
        current_row += 1

        main_start_row = current_row

        # 3a) Main shift satırları (start-bazlı gruplama)
        for start in sorted_starts:
            for day_idx in range(7):
                col_offset = day_idx * COLS_PER_DAY
                cell = ws.cell(row=current_row, column=col_offset + 1,
                               value=labels[start])
                cell.alignment = center
                cell.border = border

                day_start_counts = counts.get(start, {}).get(day_idx, {})
                vals = _row_values(day_start_counts)
                for j, v in enumerate(vals):
                    cell = ws.cell(row=current_row, column=col_offset + j + 2,
                                   value=v)
                    cell.alignment = center
                    cell.border = border
            current_row += 1

        # 3b) PT shift satırları (main_start_row'dan başlar, bağımsız büyür)
        if pt_sorted:
            pt_row = main_start_row
            for (s_start, s_end) in pt_sorted:
                label = f"{s_start}-{s_end}"
                cell = ws.cell(row=pt_row, column=PT_SECTION_START, value=label)
                cell.alignment = center
                cell.border = border
                cmt_v = pt_counts.get((s_start, s_end), {}).get(5, 0)
                pzr_v = pt_counts.get((s_start, s_end), {}).get(6, 0)
                for j, v in enumerate([cmt_v, pzr_v]):
                    cell = ws.cell(row=pt_row, column=PT_SECTION_START + 1 + j,
                                   value=v)
                    cell.alignment = center
                    cell.border = border
                pt_row += 1
            pt_toplam_row = pt_row
        else:
            pt_toplam_row = main_start_row   # PT yok, etkisiz

        # 4) Main toplam satırı
        main_toplam_row = current_row
        for day_idx in range(7):
            col_offset = day_idx * COLS_PER_DAY
            cell = ws.cell(row=main_toplam_row, column=col_offset + 1, value='Toplam')
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = center
            cell.border = border

            totals = [0, 0, 0, 0, 0, 0, 0]
            for start in sorted_starts:
                day_start_counts = counts.get(start, {}).get(day_idx, {})
                vals = _row_values(day_start_counts)
                for j in range(7):
                    totals[j] += vals[j]
            for j, v in enumerate(totals):
                cell = ws.cell(row=main_toplam_row, column=col_offset + j + 2, value=v)
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = center
                cell.border = border
        current_row += 1

        # 4b) PT toplam (kendi satırında — main toplam'a hizalı değil)
        if pt_sorted:
            cell = ws.cell(row=pt_toplam_row, column=PT_SECTION_START,
                           value='Toplam')
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = center
            cell.border = border
            cmt_total = sum(pt_counts.get(s, {}).get(5, 0) for s in pt_sorted)
            pzr_total = sum(pt_counts.get(s, {}).get(6, 0) for s in pt_sorted)
            for j, v in enumerate([cmt_total, pzr_total]):
                cell = ws.cell(row=pt_toplam_row, column=PT_SECTION_START + 1 + j,
                               value=v)
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = center
                cell.border = border

        # Bir sonraki haftaya geçmeden önce: main + PT'nin son satırının max'i
        current_row = max(current_row, pt_toplam_row + 1) + 2

    # 5) Kolon genişlikleri — ana 7 günlük bloklar
    for day_idx in range(7):
        col_offset = day_idx * COLS_PER_DAY
        ws.column_dimensions[
            ws.cell(row=2, column=col_offset + 1).column_letter].width = 12
        for j in range(1, COLS_PER_DAY):
            col_letter = ws.cell(row=2, column=col_offset + j + 1).column_letter
            ws.column_dimensions[col_letter].width = 11

    # PT section kolon genişlikleri
    for j in range(PT_SECTION_COLS):
        col_letter = ws.cell(row=2, column=PT_SECTION_START + j).column_letter
        ws.column_dimensions[col_letter].width = 12 if j == 0 else 9

    # ---- Sheet 2: Haftasonu Bütçe ----
    ws2 = wb.create_sheet('Bütçe')
    weekend_days = [(d, pd.to_datetime(d).weekday()) for d in all_dates
                    if pd.to_datetime(d).weekday() >= 5]

    # Başlık satırı
    headers = ['Kuyruk']
    for date_str, wd in weekend_days:
        label = ('Cmt' if wd == 5 else 'Pzr') + date_str[-2:]
        headers.append(label)
    headers += ['Toplam', 'Bütçe', 'Fark', 'Durum']
    for j, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    row_idx = 2
    for queue, budget in weekend_budget.items():
        ws2.cell(row=row_idx, column=1, value=queue).border = border
        total = 0
        for j, (date_str, _) in enumerate(weekend_days, 2):
            r = results.get(date_str)
            if r and queue in r and r[queue]:
                mi = r[queue]['mip_info']
                inhouse_ft = max(0, mi.get('total_inhouse_kisi', 0)
                                  - mi.get('total_part_time_kisi', 0))
            else:
                inhouse_ft = 0
            cell = ws2.cell(row=row_idx, column=j, value=inhouse_ft)
            cell.alignment = center
            cell.border = border
            total += inhouse_ft
        diff = total - budget
        status = 'AŞIM' if diff > 0 else 'OK'
        last_cols = [total, budget, diff, status]
        start_j = 2 + len(weekend_days)
        for k, v in enumerate(last_cols):
            cell = ws2.cell(row=row_idx, column=start_j + k, value=v)
            cell.alignment = center
            cell.border = border
            if k == 3:  # Durum
                cell.font = Font(bold=True,
                                  color='C00000' if status == 'AŞIM' else '006100')
        row_idx += 1

    ws2.column_dimensions['A'].width = 12
    for j in range(2, len(headers) + 1):
        col_letter = ws2.cell(row=1, column=j).column_letter
        ws2.column_dimensions[col_letter].width = 9

    # ---- Kaydet ----
    wb.save(output_path)

    n_rows_plan = ws.max_row
    print(f"✓ Excel kaydedildi: {output_path}")
    print(f"  Sheet 'Vardiya Planı': {len(weeks)} hafta, {n_rows_plan} satır")
    print(f"  Sheet 'Bütçe': {len(weekend_budget)} kuyruk, "
          f"{len(weekend_days)} hafta sonu günü")
    return output_path


export_monthly_plan_to_excel(
    results=results,
    year=YEAR,
    month=MONTH,
    weekend_budget=WEEKEND_BUDGET,
    queues=('kitle', 'gold', 'kurumsal'),
)
